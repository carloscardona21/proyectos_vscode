from flask import Flask, render_template, request, redirect, url_for, jsonify, Response
import json
import os
from datetime import datetime, timedelta
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Archivos para almacenar datos
EQUIPOS_FILE = 'equipos.json'
PRESTAMOS_FILE = 'prestamos.json'

def inicializar_archivos():
    """Inicializa los archivos JSON si no existen o están vacíos"""
    archivos = [EQUIPOS_FILE, PRESTAMOS_FILE]
    for archivo in archivos:
        if not os.path.exists(archivo):
            # Si el archivo no existe, crearlo con un array vacío
            with open(archivo, 'w', encoding='utf-8') as f:
                json.dump([], f, indent=4, ensure_ascii=False)
        else:
            # Si el archivo existe pero está vacío, escribir un array vacío
            with open(archivo, 'r', encoding='utf-8') as f:
                contenido = f.read().strip()
                if not contenido:
                    with open(archivo, 'w', encoding='utf-8') as f2:
                        json.dump([], f2, indent=4, ensure_ascii=False)

def cargar_datos(archivo):
    """Carga datos desde un archivo JSON manejando archivos vacíos o corruptos"""
    try:
        if os.path.exists(archivo):
            with open(archivo, 'r', encoding='utf-8') as f:
                contenido = f.read().strip()
                if contenido:
                    return json.loads(contenido)
                else:
                    return []
        else:
            return []
    except json.JSONDecodeError:
        # Si hay error al decodificar, retornar lista vacía y recrear archivo
        print(f"Error al decodificar {archivo}, recreando archivo...")
        with open(archivo, 'w', encoding='utf-8') as f:
            json.dump([], f, indent=4, ensure_ascii=False)
        return []

def guardar_datos(archivo, datos):
    """Guarda datos en un archivo JSON"""
    try:
        with open(archivo, 'w', encoding='utf-8') as f:
            json.dump(datos, f, indent=4, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Error al guardar en {archivo}: {e}")
        return False

# Inicializar archivos al inicio
inicializar_archivos()

@app.route('/')
def index():
    equipos = cargar_datos(EQUIPOS_FILE)
    return render_template('index.html', equipos=equipos)

@app.route('/agregar_equipo', methods=['GET', 'POST'])
def agregar_equipo():
    if request.method == 'POST':
        equipos = cargar_datos(EQUIPOS_FILE)
        nuevo_equipo = {
            'id': len(equipos) + 1,
            'nombre': request.form['nombre'],
            'tipo': request.form['tipo'],
            'marca': request.form['marca'],
            'modelo': request.form['modelo'],
            'serial': request.form['serial'],
            'estado': 'Disponible',
            'fecha_registro': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        equipos.append(nuevo_equipo)
        if guardar_datos(EQUIPOS_FILE, equipos):
            return redirect(url_for('index'))
        else:
            return "Error al guardar el equipo", 500
    return render_template('agregar_equipo.html')

@app.route('/prestar_equipo/<int:equipo_id>', methods=['POST'])
def prestar_equipo(equipo_id):
    equipos = cargar_datos(EQUIPOS_FILE)
    prestamos = cargar_datos(PRESTAMOS_FILE)
    
    # Buscar el equipo
    equipo = next((e for e in equipos if e['id'] == equipo_id), None)
    
    if equipo and equipo['estado'] == 'Disponible':
        # Actualizar estado del equipo
        equipo['estado'] = 'Prestado'
        
        # Crear registro de préstamo
        nuevo_prestamo = {
            'id': len(prestamos) + 1,
            'equipo_id': equipo_id,
            'equipo_nombre': equipo['nombre'],
            'usuario': request.form['usuario'],
            'departamento': request.form['departamento'],
            'fecha_prestamo': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'fecha_devolucion': None,
            'estado': 'Activo'
        }
        prestamos.append(nuevo_prestamo)
        
        if guardar_datos(EQUIPOS_FILE, equipos) and guardar_datos(PRESTAMOS_FILE, prestamos):
            return redirect(url_for('index'))
    
    return redirect(url_for('index'))

@app.route('/devolver_equipo/<int:prestamo_id>', methods=['POST'])
def devolver_equipo(prestamo_id):
    equipos = cargar_datos(EQUIPOS_FILE)
    prestamos = cargar_datos(PRESTAMOS_FILE)
    
    # Buscar el préstamo
    prestamo = next((p for p in prestamos if p['id'] == prestamo_id), None)
    
    if prestamo and prestamo['estado'] == 'Activo':
        # Actualizar estado del préstamo
        prestamo['estado'] = 'Devuelto'
        prestamo['fecha_devolucion'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Actualizar estado del equipo
        equipo = next((e for e in equipos if e['id'] == prestamo['equipo_id']), None)
        if equipo:
            equipo['estado'] = 'Disponible'
        
        if guardar_datos(EQUIPOS_FILE, equipos) and guardar_datos(PRESTAMOS_FILE, prestamos):
            return redirect(url_for('ver_prestamos'))
    
    return redirect(url_for('ver_prestamos'))

@app.route('/prestamos')
def ver_prestamos():
    try:
        prestamos = cargar_datos(PRESTAMOS_FILE)
        # Verificar que prestamos sea una lista
        if not isinstance(prestamos, list):
            prestamos = []
        return render_template('prestamos.html', prestamos=prestamos)
    except Exception as e:
        print(f"Error al cargar préstamos: {e}")
        return render_template('prestamos.html', prestamos=[])

@app.route('/api/equipos')
def api_equipos():
    equipos = cargar_datos(EQUIPOS_FILE)
    return jsonify(equipos)

# ============= FUNCIONES DE EXPORTACIÓN =============

@app.route('/exportar_csv/equipos')
def exportar_csv_equipos():
    """Exporta la lista de equipos a CSV"""
    equipos = cargar_datos(EQUIPOS_FILE)
    
    # Crear un buffer de texto
    output = io.StringIO()
    writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    
    # Escribir encabezados
    writer.writerow(['ID', 'Nombre', 'Tipo', 'Marca', 'Modelo', 'Serial', 'Estado', 'Fecha Registro'])
    
    # Escribir datos
    for equipo in equipos:
        writer.writerow([
            equipo['id'],
            equipo['nombre'],
            equipo['tipo'],
            equipo['marca'],
            equipo['modelo'],
            equipo['serial'],
            equipo['estado'],
            equipo['fecha_registro']
        ])
    
    # Configurar la respuesta
    output.seek(0)
    response = Response(output.getvalue(), mimetype='text/csv')
    response.headers['Content-Disposition'] = 'attachment; filename=equipos.csv'
    return response

@app.route('/exportar_excel/equipos')
def exportar_excel_equipos():
    """Exporta la lista de equipos a Excel"""
    equipos = cargar_datos(EQUIPOS_FILE)
    
    # Crear libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipos Tecnológicos"
    
    # Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir encabezados
    headers = ['ID', 'Nombre', 'Tipo', 'Marca', 'Modelo', 'Serial', 'Estado', 'Fecha Registro']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Escribir datos
    for row, equipo in enumerate(equipos, 2):
        ws.cell(row=row, column=1, value=equipo['id'])
        ws.cell(row=row, column=2, value=equipo['nombre'])
        ws.cell(row=row, column=3, value=equipo['tipo'])
        ws.cell(row=row, column=4, value=equipo['marca'])
        ws.cell(row=row, column=5, value=equipo['modelo'])
        ws.cell(row=row, column=6, value=equipo['serial'])
        
        # Color según estado
        estado_cell = ws.cell(row=row, column=7, value=equipo['estado'])
        if equipo['estado'] == 'Disponible':
            estado_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif equipo['estado'] == 'Prestado':
            estado_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws.cell(row=row, column=8, value=equipo['fecha_registro'])
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Congelar el panel superior
    ws.freeze_panes = 'A2'
    
    # Guardar en buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Configurar respuesta
    response = Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response.headers['Content-Disposition'] = 'attachment; filename=equipos.xlsx'
    return response

@app.route('/exportar_csv/prestamos')
def exportar_csv_prestamos():
    """Exporta el historial de préstamos a CSV"""
    prestamos = cargar_datos(PRESTAMOS_FILE)
    
    # Crear un buffer de texto
    output = io.StringIO()
    writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    
    # Escribir encabezados
    writer.writerow(['ID', 'Equipo', 'Usuario', 'Departamento', 'Fecha Préstamo', 'Fecha Devolución', 'Estado'])
    
    # Escribir datos
    for prestamo in prestamos:
        writer.writerow([
            prestamo['id'],
            prestamo['equipo_nombre'],
            prestamo['usuario'],
            prestamo['departamento'],
            prestamo['fecha_prestamo'],
            prestamo['fecha_devolucion'] if prestamo['fecha_devolucion'] else 'Pendiente',
            prestamo['estado']
        ])
    
    # Configurar la respuesta
    output.seek(0)
    response = Response(output.getvalue(), mimetype='text/csv')
    response.headers['Content-Disposition'] = 'attachment; filename=prestamos.csv'
    return response

@app.route('/exportar_excel/prestamos')
def exportar_excel_prestamos():
    """Exporta el historial de préstamos a Excel"""
    prestamos = cargar_datos(PRESTAMOS_FILE)
    
    # Crear libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Préstamos"
    
    # Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir encabezados
    headers = ['ID', 'Equipo', 'Usuario', 'Departamento', 'Fecha Préstamo', 'Fecha Devolución', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Escribir datos
    for row, prestamo in enumerate(prestamos, 2):
        ws.cell(row=row, column=1, value=prestamo['id'])
        ws.cell(row=row, column=2, value=prestamo['equipo_nombre'])
        ws.cell(row=row, column=3, value=prestamo['usuario'])
        ws.cell(row=row, column=4, value=prestamo['departamento'])
        ws.cell(row=row, column=5, value=prestamo['fecha_prestamo'])
        ws.cell(row=row, column=6, value=prestamo['fecha_devolucion'] if prestamo['fecha_devolucion'] else 'Pendiente')
        
        # Color según estado
        estado_cell = ws.cell(row=row, column=7, value=prestamo['estado'])
        if prestamo['estado'] == 'Activo':
            estado_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif prestamo['estado'] == 'Devuelto':
            estado_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Congelar el panel superior
    ws.freeze_panes = 'A2'
    
    # Guardar en buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Configurar respuesta
    response = Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response.headers['Content-Disposition'] = 'attachment; filename=prestamos.xlsx'
    return response

@app.route('/exportar_excel/completo')
def exportar_excel_completo():
    """Exporta un informe completo con múltiples hojas"""
    equipos = cargar_datos(EQUIPOS_FILE)
    prestamos = cargar_datos(PRESTAMOS_FILE)
    
    wb = Workbook()
    
    # Hoja de equipos
    ws_equipos = wb.active
    ws_equipos.title = "Equipos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados equipos
    headers_equipos = ['ID', 'Nombre', 'Tipo', 'Marca', 'Modelo', 'Serial', 'Estado', 'Fecha Registro']
    for col, header in enumerate(headers_equipos, 1):
        cell = ws_equipos.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos equipos
    for row, equipo in enumerate(equipos, 2):
        ws_equipos.cell(row=row, column=1, value=equipo['id'])
        ws_equipos.cell(row=row, column=2, value=equipo['nombre'])
        ws_equipos.cell(row=row, column=3, value=equipo['tipo'])
        ws_equipos.cell(row=row, column=4, value=equipo['marca'])
        ws_equipos.cell(row=row, column=5, value=equipo['modelo'])
        ws_equipos.cell(row=row, column=6, value=equipo['serial'])
        
        # Color según estado
        estado_cell = ws_equipos.cell(row=row, column=7, value=equipo['estado'])
        if equipo['estado'] == 'Disponible':
            estado_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif equipo['estado'] == 'Prestado':
            estado_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws_equipos.cell(row=row, column=8, value=equipo['fecha_registro'])
    
    # Hoja de préstamos
    ws_prestamos = wb.create_sheet("Préstamos")
    
    # Encabezados préstamos
    headers_prestamos = ['ID', 'Equipo', 'Usuario', 'Departamento', 'Fecha Préstamo', 'Fecha Devolución', 'Estado']
    for col, header in enumerate(headers_prestamos, 1):
        cell = ws_prestamos.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos préstamos
    for row, prestamo in enumerate(prestamos, 2):
        ws_prestamos.cell(row=row, column=1, value=prestamo['id'])
        ws_prestamos.cell(row=row, column=2, value=prestamo['equipo_nombre'])
        ws_prestamos.cell(row=row, column=3, value=prestamo['usuario'])
        ws_prestamos.cell(row=row, column=4, value=prestamo['departamento'])
        ws_prestamos.cell(row=row, column=5, value=prestamo['fecha_prestamo'])
        ws_prestamos.cell(row=row, column=6, value=prestamo['fecha_devolucion'] if prestamo['fecha_devolucion'] else 'Pendiente')
        
        # Color según estado
        estado_cell = ws_prestamos.cell(row=row, column=7, value=prestamo['estado'])
        if prestamo['estado'] == 'Activo':
            estado_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif prestamo['estado'] == 'Devuelto':
            estado_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Hoja de estadísticas
    ws_stats = wb.create_sheet("Estadísticas")
    
    # Calcular estadísticas
    total_equipos = len(equipos)
    disponibles = len([e for e in equipos if e['estado'] == 'Disponible'])
    prestados = len([e for e in equipos if e['estado'] == 'Prestado'])
    prestamos_activos = len([p for p in prestamos if p['estado'] == 'Activo'])
    prestamos_devueltos = len([p for p in prestamos if p['estado'] == 'Devuelto'])
    
    # Escribir estadísticas con formato
    stats_data = [
        ['INFORME COMPLETO DEL SISTEMA', ''],
        ['Fecha de generación:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['', ''],
        ['ESTADÍSTICAS DE EQUIPOS', ''],
        ['Total Equipos', total_equipos],
        ['Equipos Disponibles', disponibles],
        ['Equipos Prestados', prestados],
        ['Tasa de disponibilidad', f'{(disponibles/total_equipos*100 if total_equipos > 0 else 0):.1f}%'],
        ['', ''],
        ['ESTADÍSTICAS DE PRÉSTAMOS', ''],
        ['Préstamos Activos', prestamos_activos],
        ['Préstamos Devueltos', prestamos_devueltos],
        ['Total Préstamos', len(prestamos)],
    ]
    
    # Aplicar formato a las estadísticas
    for row, (key, value) in enumerate(stats_data, 1):
        cell_key = ws_stats.cell(row=row, column=1, value=key)
        cell_value = ws_stats.cell(row=row, column=2, value=value)
        
        # Formato para títulos
        if 'INFORME' in key or 'ESTADÍSTICAS' in key:
            cell_key.font = Font(bold=True, size=12, color="2C3E50")
            if cell_value.value:
                cell_value.font = Font(bold=True, size=12, color="2C3E50")
        elif key and not value:
            pass
    
    # Ajustar anchos de columnas para todas las hojas
    for ws in [ws_equipos, ws_prestamos, ws_stats]:
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Congelar paneles
    ws_equipos.freeze_panes = 'A2'
    ws_prestamos.freeze_panes = 'A2'
    
    # Guardar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response.headers['Content-Disposition'] = 'attachment; filename=informe_completo.xlsx'
    return response

if __name__ == '__main__':
    # Asegurar que los archivos existan antes de iniciar
    inicializar_archivos()
    app.run(debug=True)